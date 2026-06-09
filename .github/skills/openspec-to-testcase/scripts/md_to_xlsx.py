#!/usr/bin/env python3
"""
md_to_xlsx.py
Converts tests/ markdown files (test-plan.md, rtm.md, tests/<cap>/test-cases.md)
into a structured Excel workbook for test execution.

Usage:
    python scripts/md_to_xlsx.py --tests-dir tests/ --output <ProjectName>_Tests_20260402.xlsx

    <ProjectName>: derive from the 'project:' field in tests/test-plan.md frontmatter,
    or use the repository folder name in TitleCase if the field is absent.

Requirements:
    pip install openpyxl pyyaml
"""

import argparse
import re
import sys
from datetime import date
from pathlib import Path

try:
    import yaml
except ImportError:
    yaml = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colour palette ─────────────────────────────────────────────────────────────
C = {
    "purple_dark":  "3C3489", "purple_mid":   "534AB7", "purple_light": "EEEDFE",
    "teal_dark":    "085041", "teal_mid":     "1D9E75", "teal_light":   "E1F5EE",
    "amber_dark":   "633806", "amber_mid":    "BA7517", "amber_light":  "FAEEDA",
    "gray_dark":    "444441", "gray_mid":     "888780", "gray_light":   "F1EFE8",
    "red_light":    "FCEBEB", "red_mid":      "E24B4A",
    "green_light":  "EAF3DE", "green_mid":    "639922",
    "white":        "FFFFFF", "row_alt":      "F7F6FF",
}

def _fill(h): return PatternFill("solid", start_color=h, end_color=h)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _font(color="2C2C2A", sz=10, bold=False): return Font(name="Arial", size=sz, bold=bold, color=color)
def _hfont(color="FFFFFF", sz=11): return Font(name="Arial", size=sz, bold=True, color=color)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def _title_row(ws, text, color, ncols, row=1, sz=13):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(color); c.font = _hfont(sz=sz); c.alignment = _center()
    ws.row_dimensions[row].height = 28

def _hrow(ws, row, headers_widths, color):
    for ci, (h, w) in enumerate(headers_widths, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(color); c.font = _hfont(sz=10)
        c.alignment = _center(); c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 28

def _drow(ws, row, vals, alt=False, heights=None):
    bg = C["row_alt"] if alt else C["white"]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _left()
    if heights:
        ws.row_dimensions[row].height = heights

# ── Markdown parsers ────────────────────────────────────────────────────────────

def _strip_yaml(text):
    """Remove YAML frontmatter block and return (meta_dict, body)."""
    meta = {}
    if text.startswith("---"):
        end = text.find("\n---", 3)
        if end != -1:
            raw = text[3:end].strip()
            if yaml:
                try: meta = yaml.safe_load(raw) or {}
                except: pass
            text = text[end+4:].strip()
    return meta, text

def parse_test_cases(path: Path) -> list[dict]:
    """Parse a test-cases.md file into a list of TC dicts."""
    text = path.read_text(encoding="utf-8")
    _, body = _strip_yaml(text)
    tcs = []
    # Split on ## TC- headings
    blocks = re.split(r'\n(?=## TC-)', body)
    for block in blocks:
        m = re.match(r'## (TC-[\w-]+)\s+[—–-]+\s+(.+)', block)
        if not m:
            continue
        tc_id = m.group(1).strip()
        tc_name = m.group(2).strip()

        def _field(label):
            pat = rf'\|\s*\*\*{label}\*\*\s*\|\s*(.+?)\s*\|'
            fm = re.search(pat, block)
            return fm.group(1).strip() if fm else ""

        def _section(label):
            # Stop only at a line that starts a new section header (**Word...:**)
            # or at end of block. This prevents truncation on inline bold text inside
            # a section body (e.g., "**Note:**" embedded in a step description).
            pat = rf'\*\*{re.escape(label)}:\*\*\n(.*?)(?=\n\*\*[A-Z][^*\n]*:\*\*|\Z)'
            sm = re.search(pat, block, re.DOTALL)
            return sm.group(1).strip() if sm else ""

        req_ids   = _field("REQ ID\\(s\\)")
        tc_type   = _field("Type")
        priority  = _field("Priority")
        automation= _field("Automation")
        status    = _field("Status")
        desc      = _section("Description")
        precond   = _section("Preconditions")
        steps     = _section("Steps to Test")
        test_data = _section("Test Data")
        expected  = _section("Expected Result")
        actual    = _section("Actual Result")
        notes     = _section("Notes")

        # Clean up actual result placeholder
        if "_[Fill at execution time]_" in actual:
            actual = ""

        tcs.append({
            "tc_id": tc_id, "req_ids": req_ids, "name": tc_name,
            "type": tc_type, "priority": priority, "automation": automation,
            "status": status, "description": desc, "preconditions": precond,
            "steps": steps, "test_data": test_data, "expected": expected,
            "actual": actual, "notes": notes,
            "capability": path.parent.name,
        })
    return tcs

def parse_rtm(path: Path) -> tuple[list[dict], list[dict]]:
    """Parse rtm.md → (req_rows, tc_rows)."""
    text = path.read_text(encoding="utf-8")
    _, body = _strip_yaml(text)
    req_rows, tc_rows = [], []

    # Traceability Table
    in_trace = False
    for line in body.splitlines():
        if "## Traceability Table" in line:
            in_trace = True; continue
        if in_trace and line.startswith("## "):
            in_trace = False
        if in_trace and line.startswith("|"):
            parts = [p.strip() for p in line.strip().strip("|").split("|")]
            # Skip header row, separator row, or empty
            if not parts or parts[0] in ("REQ ID", "---", "") or all(set(p) <= {'-'} for p in parts if p):
                continue
            if len(parts) >= 5 and parts[0].startswith("REQ-"):
                req_rows.append({
                    "req_id": parts[0], "area": parts[1], "description": parts[2],
                    "priority": parts[3], "coverage": parts[4],
                    "test_cases": parts[5] if len(parts) > 5 else "",
                })

    # Reverse Index
    in_rev = False
    for line in body.splitlines():
        if "## Reverse Index" in line:
            in_rev = True; continue
        if in_rev and line.startswith("## "):
            in_rev = False
        if in_rev and line.startswith("|"):
            parts = [p.strip() for p in line.strip().strip("|").split("|")]
            if not parts or parts[0] in ("TC ID", "---", "") or all(set(p) <= {'-'} for p in parts if p):
                continue
            if len(parts) >= 5 and parts[0].startswith("TC-"):
                tc_rows.append({
                    "tc_id": parts[0], "type": parts[1], "req_ids": parts[2],
                    "capability": parts[3], "status": parts[4],
                })
    return req_rows, tc_rows

def parse_test_plan(path: Path) -> list[tuple[str, str]]:
    """Parse test-plan.md into (label, value) pairs for the plan sheet."""
    text = path.read_text(encoding="utf-8")
    _, body = _strip_yaml(text)
    rows = []
    current_section = None
    for line in body.splitlines():
        if line.startswith("## "):
            current_section = line.lstrip("# ").strip()
            rows.append((f"── {current_section.upper()} ──", ""))
        elif line.startswith("| ") and not line.startswith("|---") and not line.startswith("| ---"):
            parts = [p.strip() for p in line.strip().strip("|").split("|")]
            if len(parts) == 2 and parts[0] and parts[1]:
                label = parts[0].replace("**","").strip()
                value = parts[1].replace("**","").strip()
                if label and not label.startswith("Field") and not label.startswith("---"):
                    rows.append((label, value))
        elif line.startswith("**") and ":" in line:
            label, _, value = line.partition(":")
            rows.append((label.replace("**","").strip(), value.strip()))
        elif line.startswith("- ") and current_section:
            rows.append(("  •", line.lstrip("- ")))
    return rows

# ── Excel sheet builders ────────────────────────────────────────────────────────

def build_readme(wb):
    ws = wb.active; ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    _title_row(ws, "Test Management Workbook — Generated from Markdown", C["purple_dark"], 2)
    rows = [
        ("SOURCE OF TRUTH", ""),
        ("Markdown files", "tests/ folder in source control — edit here"),
        ("This workbook", "Generated execution artifact — do NOT edit as source"),
        ("Regenerate", "python scripts/md_to_xlsx.py --tests-dir tests/ --output <name>.xlsx"),
        ("", ""),
        ("SHEETS", ""),
        ("Test_Plan", "Project-level test plan from tests/test-plan.md"),
        ("Test_Cases", "All TCs from tests/<capability>/test-cases.md"),
        ("RTM", "Requirements traceability from tests/rtm.md"),
        ("", ""),
        ("TC TYPE CODES", ""),
        ("P — Positive", "Happy path — valid input, expected success"),
        ("N — Negative", "Invalid input / error / resource failure"),
        ("E — Edge", "Boundary, concurrency, empty set, low-resource"),
        ("", ""),
        ("STATUS", ""),
        ("Not Run", "Not yet executed"),
        ("Pass", "Actual result matched expected"),
        ("Fail", "Actual result did NOT match expected"),
        ("Blocked", "Cannot run — dependency or environment issue"),
        ("", ""),
        ("COVERAGE (RTM)", ""),
        ("Full", "Has both P and N test cases"),
        ("Partial", "P only or N only"),
        ("None", "No test cases linked yet — gap!"),
        ("", ""),
        (f"Generated", str(date.today())),
    ]
    section_labels = {"SOURCE OF TRUTH", "SHEETS", "TC TYPE CODES", "STATUS", "COVERAGE (RTM)"}
    for i, (label, value) in enumerate(rows, 2):
        ws.row_dimensions[i].height = 18
        ca = ws.cell(row=i, column=1, value=label)
        cb = ws.cell(row=i, column=2, value=value)
        if label in section_labels:
            for c in (ca, cb):
                c.fill = _fill(C["gray_light"])
                c.font = Font(name="Arial", size=10, bold=True, color=C["gray_dark"])
        elif label:
            ca.font = Font(name="Arial", size=10, bold=True, color=C["purple_mid"])
            cb.font = _font()
        for c in (ca, cb):
            c.border = _border(); c.alignment = _left()

def build_plan(wb, plan_rows):
    ws = wb.create_sheet("Test_Plan")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    _title_row(ws, "Test Plan", C["teal_dark"], 2)
    section_marker = re.compile(r'^──\s+(.+)\s+──$')
    for i, (label, value) in enumerate(plan_rows, 2):
        ws.row_dimensions[i].height = 20 if "\n" not in value else 40
        ca = ws.cell(row=i, column=1, value=label)
        cb = ws.cell(row=i, column=2, value=value)
        if section_marker.match(label):
            for c in (ca, cb):
                c.fill = _fill(C["teal_mid"]); c.font = _hfont(sz=10)
        elif label == "":
            for c in (ca, cb): c.fill = _fill(C["white"])
        else:
            ca.fill = _fill(C["teal_light"])
            ca.font = Font(name="Arial", size=10, bold=True, color=C["teal_dark"])
            cb.font = _font()
        for c in (ca, cb): c.border = _border(); c.alignment = _left()

TYPE_FILL  = {"P": (C["teal_light"],  C["teal_dark"]),
              "N": (C["red_light"],   C["red_mid"]),
              "E": (C["amber_light"], C["amber_dark"])}
STATUS_FILL= {"Pass": C["green_light"], "Fail": C["red_light"],
              "Blocked": C["amber_light"], "Not Run": C["gray_light"]}

def build_tcs(wb, all_tcs):
    ws = wb.create_sheet("Test_Cases")
    ws.sheet_view.showGridLines = False; ws.freeze_panes = "A3"
    headers = [
        ("TC ID",14), ("REQ ID(s)",18), ("Capability",16), ("Test Case Name",36),
        ("Type",8), ("Priority",10), ("Description",40), ("Preconditions",36),
        ("Steps to Test",52), ("Test Data",28), ("Expected Result",40),
        ("Actual Result",40), ("Status",12), ("Automation",12), ("Notes",22),
    ]
    ncols = len(headers)
    _title_row(ws, "Test Cases", C["amber_dark"], ncols)
    _hrow(ws, 2, headers, C["amber_mid"])

    dv_type   = DataValidation(type="list", formula1='"P,N,E"',                        allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked"',     allow_blank=True)
    dv_auto   = DataValidation(type="list", formula1='"Manual,Automated,Both"',         allow_blank=True)
    for dv, col in [(dv_type,"E"), (dv_status,"M"), (dv_auto,"N")]:
        dv.sqref = f"{col}3:{col}2000"; ws.add_data_validation(dv)

    for ri, tc in enumerate(all_tcs, 3):
        alt = ri % 2 == 0
        vals = [tc["tc_id"], tc["req_ids"], tc["capability"], tc["name"],
                tc["type"], tc["priority"], tc["description"], tc["preconditions"],
                tc["steps"], tc["test_data"], tc["expected"], tc["actual"],
                tc["status"], tc["automation"], tc["notes"]]
        _drow(ws, ri, vals, alt, heights=72)
        # TC ID styling
        c = ws.cell(row=ri, column=1)
        c.fill = _fill(C["purple_light"])
        c.font = Font(name="Arial", size=10, bold=True, color=C["purple_dark"])
        # REQ ID styling
        c = ws.cell(row=ri, column=2)
        c.fill = _fill(C["teal_light"])
        c.font = Font(name="Arial", size=10, bold=True, color=C["teal_dark"])
        # Type colouring
        tc_type = tc["type"]
        if tc_type in TYPE_FILL:
            bg, fg = TYPE_FILL[tc_type]
            c = ws.cell(row=ri, column=5)
            c.fill = _fill(bg); c.font = Font(name="Arial", size=10, bold=True, color=fg)
            c.alignment = _center()
        # Status colouring
        status = tc["status"]
        if status in STATUS_FILL:
            c = ws.cell(row=ri, column=13)
            c.fill = _fill(STATUS_FILL[status]); c.alignment = _center()

def build_rtm(wb, req_rows, tc_rows):
    ws = wb.create_sheet("RTM")
    ws.sheet_view.showGridLines = False; ws.freeze_panes = "A3"
    # Summary stats
    total_req = len(req_rows)
    cov_counts = {"Full": 0, "Partial": 0, "None": 0}
    for r in req_rows: cov_counts[r.get("coverage", "None")] = cov_counts.get(r.get("coverage","None"),0)+1

    ncols = 6
    _title_row(ws, "Requirements Traceability Matrix (RTM)", C["purple_dark"], ncols)
    # Summary block
    summary = [
        f"Requirements: {total_req}  |  Full: {cov_counts.get('Full',0)}  |  "
        f"Partial: {cov_counts.get('Partial',0)}  |  None: {cov_counts.get('None',0)}  |  "
        f"Test Cases: {len(tc_rows)}"
    ]
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws.cell(row=2, column=1, value=summary[0])
    c.fill = _fill(C["purple_light"]); c.font = Font(name="Arial", size=10, bold=True, color=C["purple_dark"])
    c.alignment = _center(); ws.row_dimensions[2].height = 20

    req_headers = [("REQ ID",18),("Area",12),("Description",52),("Priority",10),("Coverage",12),("Test Cases",50)]
    _hrow(ws, 3, req_headers, C["purple_mid"])

    PRIORITY_BG = {"High": C["red_light"], "Medium": C["amber_light"], "Low": C["green_light"]}
    COVERAGE_BG = {"Full": C["green_light"], "Partial": C["amber_light"], "None": C["red_light"]}

    for ri, req in enumerate(req_rows, 4):
        alt = ri % 2 == 0
        vals = [req["req_id"], req["area"], req["description"], req["priority"], req["coverage"], req["test_cases"]]
        _drow(ws, ri, vals, alt, heights=32)
        # REQ ID
        c = ws.cell(row=ri, column=1)
        c.fill = _fill(C["purple_light"]); c.font = Font(name="Arial", size=10, bold=True, color=C["purple_dark"])
        # Priority
        c = ws.cell(row=ri, column=4)
        c.fill = _fill(PRIORITY_BG.get(req["priority"], C["white"]))
        c.alignment = _center()
        # Coverage
        cov = req["coverage"]
        c = ws.cell(row=ri, column=5)
        c.fill = _fill(COVERAGE_BG.get(cov, C["white"]))
        c.font = Font(name="Arial", size=10, bold=True, color=C["gray_dark"])
        c.alignment = _center()

    # Reverse Index
    rev_start = len(req_rows) + 6
    ws.merge_cells(f"A{rev_start}:{get_column_letter(ncols)}{rev_start}")
    c = ws.cell(row=rev_start, column=1, value="Reverse Index — TC to REQ (Backward Traceability)")
    c.fill = _fill(C["teal_mid"]); c.font = _hfont(sz=11); c.alignment = _center()
    ws.row_dimensions[rev_start].height = 24

    rev_headers = [("TC ID",16),("Type",8),("REQ ID(s)",20),("Capability",18),("Status",12),("",6)]
    _hrow(ws, rev_start+1, rev_headers, C["teal_mid"])

    for ri, tc in enumerate(tc_rows, rev_start+2):
        alt = ri % 2 == 0
        vals = [tc["tc_id"], tc["type"], tc["req_ids"], tc["capability"], tc["status"], ""]
        _drow(ws, ri, vals, alt, heights=20)
        # TC ID
        c = ws.cell(row=ri, column=1)
        c.fill = _fill(C["teal_light"]); c.font = Font(name="Arial", size=10, bold=True, color=C["teal_dark"])
        # Type
        tc_type = tc["type"]
        if tc_type in TYPE_FILL:
            bg, fg = TYPE_FILL[tc_type]
            c = ws.cell(row=ri, column=2)
            c.fill = _fill(bg); c.font = Font(name="Arial", size=10, bold=True, color=fg)
            c.alignment = _center()
        # Status
        status = tc["status"]
        if status in STATUS_FILL:
            c = ws.cell(row=ri, column=5)
            c.fill = _fill(STATUS_FILL[status]); c.alignment = _center()

# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Convert test markdown files to XLSX")
    parser.add_argument("--tests-dir", required=True, help="Path to tests/ directory")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    args = parser.parse_args()

    tests_dir = Path(args.tests_dir)
    if not tests_dir.exists():
        print(f"ERROR: tests directory not found: {tests_dir}", file=sys.stderr); sys.exit(1)

    # Collect all test-cases.md files
    all_tcs = []
    for tc_file in sorted(tests_dir.rglob("test-cases.md")):
        tcs = parse_test_cases(tc_file)
        all_tcs.extend(tcs)
        print(f"  Parsed {len(tcs)} TCs from {tc_file}")

    # RTM
    rtm_path = tests_dir / "rtm.md"
    req_rows, tc_rows = [], []
    if rtm_path.exists():
        req_rows, tc_rows = parse_rtm(rtm_path)
        print(f"  Parsed {len(req_rows)} REQs and {len(tc_rows)} TC rows from rtm.md")
    else:
        print("  WARNING: tests/rtm.md not found — RTM sheet will be empty")

    # Test plan
    plan_path = tests_dir / "test-plan.md"
    plan_rows = []
    if plan_path.exists():
        plan_rows = parse_test_plan(plan_path)
        print(f"  Parsed {len(plan_rows)} rows from test-plan.md")
    else:
        print("  WARNING: tests/test-plan.md not found — Test_Plan sheet will be empty")

    # Build workbook
    wb = Workbook()
    build_readme(wb)
    build_plan(wb, plan_rows)
    build_tcs(wb, all_tcs)
    build_rtm(wb, req_rows, tc_rows)

    out = Path(args.output)
    wb.save(out)
    print(f"\nDone → {out}")
    print(f"  Sheets: README, Test_Plan, Test_Cases ({len(all_tcs)} rows), RTM ({len(req_rows)} REQs)")

if __name__ == "__main__":
    main()
